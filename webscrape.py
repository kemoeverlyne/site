import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Open the Excel file containing the URLs
workbook = load_workbook('trial.xlsx')
worksheet = workbook.active
header_row = list(worksheet.rows)[0]
status_column = header_row[-2].column_letter
reason_column = header_row[-1].column_letter

# Loop through each URL in the Excel file
for row in worksheet.iter_rows(min_row=2):
    url = row[0].value
    print(f"Checking URL: {url}")

    # Check if the URL is valid
    try:
        response = requests.get(url)
        response.raise_for_status()
    except requests.exceptions.RequestException:
        # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
        status_cell = f"{status_column}{row[0].row}"
        worksheet[status_cell] = 'Fail'
        reason_cell = f"{reason_column}{row[0].row}"
        worksheet[reason_cell] = 'Invalid URL'
        continue

    # Send a GET request to the URL
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the dropdown menu button
    dropdown_button = soup.find('button', {'data-name': 'MAIN_NAV_TRIGGER'})

    # Check if the dropdown menu button exists
    if dropdown_button:
        print("Dropdown menu works!")
    else:
        print("Dropdown menu does not work!")
        # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
        status_cell = f"{status_column}{row[0].row}"
        worksheet[status_cell] = 'Fail'
        reason_cell = f"{reason_column}{row[0].row}"
        worksheet[reason_cell] = 'Dropdown menu not working'
        continue

    # Check if all pages with links on the home page are translated to Hindi including the homepage
    homepage_links = soup.select('a')
    for link in homepage_links:
        link_url = link.get('href')
        if link_url and not link_url.startswith('#'):
            try:
                link_response = requests.get(link_url)
                link_response.raise_for_status()
            except requests.exceptions.RequestException:
                # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
                status_cell = f"{status_column}{row[0].row}"
                worksheet[status_cell] = 'Fail'
                reason_cell = f"{reason_column}{row[0].row}"
                worksheet[reason_cell] = f"Invalid URL: {link_url}"
                continue

            link_soup = BeautifulSoup(link_response.content, 'html.parser')
            link_lang = link_soup.find('html').get('lang')
            if link_lang == 'hi':
                print(f"{link_url} passed translation check")
            else:
                print(f"{link_url} failed translation check")
                # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
                status_cell = f"{status_column}{row[0].row}"
                worksheet[status_cell] = 'Fail'
                reason_cell = f"{reason_column}{row[0].row}"
                worksheet[reason_cell] = f"Translation not correct: {link_url}"
                continue

    # Check if images are not blurred
    images = soup.find_all('img')
    for image in images:
        if 'blur' in image.get('class', []):
            print(f"{url} failed image check: {image['src']} is blurred")
            # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
            status_cell = f"{status_column}{row[0].row}"
            worksheet[status_cell] = 'Fail'
            reason_cell = f"{reason_column}{row[0].row}"
            worksheet[reason_cell] = f"Blurred image: {image['src']}"
            break
    else:
        print(f"{url} passed image check")
    # Check if the page title is correct
    title = soup.find('title')
    expected_title = row[1].value
    if title and title.string == expected_title:
        print(f"{url} passed title check")
    else:
        print(f"{url} failed title check")
    # Write 'Fail' to the status column and a reason to the reason column, and move to the next row
        status_cell = f"{status_column}{row[0].row}"
        worksheet[status_cell] = 'Fail'
        reason_cell = f"{reason_column}{row[0].row}"
        worksheet[reason_cell] = 'Incorrect page title'
        continue

# Write 'Pass' to the status column if all checks have passed
status_cell = f"{status_column}{row[0].row}"
worksheet[status_cell] = 'Pass'
workbook.save('trial2_results.xlsx')
